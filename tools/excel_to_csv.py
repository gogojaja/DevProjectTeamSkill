#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
excel_to_csv.py — 存量 .xlsx 一次性迁移为 CSV（每 Sheet 一个 CSV，UTF-8 with BOM）

规范：references/token_standard.md §3
用法：
  python excel_to_csv.py <file.xlsx> [输出目录]        # 单文件，每 Sheet 一个 CSV
  python excel_to_csv.py <dir> --recursive             # 目录递归批量转换
  python excel_to_csv.py <file.xlsx> --sheet "启动组"  # 只转指定 Sheet

输出：
  <同名>.csv 或 <输出目录>/<Sheet名>.csv（自动加 NN_ 序号前缀）
  转换后打印每个 Sheet 的行数，供与源文件比对。
"""

import argparse
import csv
import os
import sys

try:
    import openpyxl
except ImportError:
    print("需要安装 openpyxl：pip install openpyxl")
    sys.exit(1)

COL_LIMIT = 30  # token_standard §3.2-4：单 CSV 列数上限


def sheet_to_csv(wb, sheet_name, out_dir, prefix=""):
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        print(f"  [跳过] {sheet_name}: 空 Sheet")
        return None
    header = [str(c) if c is not None else "" for c in rows[0]]
    if len(header) > COL_LIMIT:
        print(f"  [警告] {sheet_name}: {len(header)} 列 > {COL_LIMIT}，建议按域拆分")
    safe = "".join(c for c in sheet_name if c not in r'\/:*?"<>|')
    out_path = os.path.join(out_dir, f"{prefix}{safe}.csv")
    with open(out_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(header)
        for r in rows[1:]:
            writer.writerow(["" if c is None else c for c in r])
    return len(rows)


def convert_file(xlsx_path, out_dir, only_sheet=None):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if only_sheet:
        sheets = [only_sheet]
    else:
        sheets = wb.sheetnames
    prefix = f"{os.path.splitext(os.path.basename(xlsx_path))[0].replace('总台账', '台账').replace('台账', '')}"
    total = 0
    print(f"转换: {xlsx_path}")
    for s in sheets:
        if s not in wb.sheetnames:
            print(f"  [错误] Sheet '{s}' 不存在")
            continue
        n = sheet_to_csv(wb, s, out_dir, prefix="")
        if n:
            total += n - 1
    print(f"  完成: {len(sheets)} Sheet, 数据行合计 {total}")
    return total


def convert_dir(dir_path, recursive=False):
    total_files = 0
    for root, dirs, files in os.walk(dir_path):
        for f in sorted(files):
            if f.lower().endswith(".xlsx"):
                out_dir = root
                convert_file(os.path.join(root, f), out_dir)
                total_files += 1
        if not recursive:
            break
    print(f"\n共转换 {total_files} 个 .xlsx 文件")


def main():
    parser = argparse.ArgumentParser(description="存量 .xlsx 迁移为 CSV（UTF-8 with BOM）")
    parser.add_argument("target", help=".xlsx 文件或目录")
    parser.add_argument("outdir", nargs="?", help="输出目录（缺省为源文件所在目录）")
    parser.add_argument("--recursive", action="store_true", help="目录递归")
    parser.add_argument("--sheet", help="只转换指定 Sheet")
    args = parser.parse_args()

    if os.path.isfile(args.target):
        out_dir = args.outdir or os.path.dirname(args.target)
        os.makedirs(out_dir, exist_ok=True)
        convert_file(args.target, out_dir, args.sheet)
    elif os.path.isdir(args.target):
        convert_dir(args.target, args.recursive)
    else:
        print(f"路径不存在: {args.target}")
        sys.exit(1)


if __name__ == "__main__":
    main()
